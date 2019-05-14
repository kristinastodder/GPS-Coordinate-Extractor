#! python3
# This is a testing document

import openpyxl, pyperclip, requests, os, bs4

from selenium import webdriver
import time

#Open Firefox browser and latlong.net
driver = webdriver.Firefox()
driver.get('http://latlong.net')

# Open the spreadsheet and copy the value
wb = openpyxl.load_workbook('California.xlsx')
sheet = wb['Sheet1']
for r in range(2, sheet.max_row + 1):
    location = sheet.cell(row=r, column=7).value
    pyperclip.copy(location)

    #Submit location
    placeElem = driver.find_element_by_id('place')
    placeElem.send_keys(location)
    placeElem.submit()
    time.sleep(5)

    #Copy lat & long to clipboard and print
    latlng = driver.find_element_by_id('latlngspan')
    pyperclip.copy(latlng.text)
    print(latlng.text)

    #Browser Back
    driver.back()
    
elif location = ""
    prin("Value is null")



